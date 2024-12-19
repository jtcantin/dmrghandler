import datetime
import json
from pathlib import Path

import h5py
import numpy as np
import openpyxl as px
import openpyxl.chart as px_chart
import pandas as pd
from openpyxl.chart.layout import Layout, ManualLayout
import uuid


def get_data_from_incomplete_processing(data_file):
    # Get DMRG energy, bond dimensions, and truncation error for each loop
    dmrg_energies = []
    bond_dimensions = []
    discarded_weights = []
    loop_cpu_times_s = []
    loop_wall_times_s = []
    num_sweeps_list = []
    final_sweep_delta_energies_list = []
    reordering_method_list = []
    reordering_method_cpu_times_s = []
    reordering_method_wall_times_s = []
    extra_dict = {
        "loop_copy_mps_cpu_time_s_list": [],
        "loop_dmrg_optimization_cpu_time_s_list": [],
        "loop_driver_initialize_system_cpu_time_s_list": [],
        "loop_generate_initial_mps_cpu_time_s_list": [],
        "loop_get_qchem_hami_mpo_cpu_time_s_list": [],
        "loop_make_driver_cpu_time_s_list": [],
        "loop_reorder_integrals_cpu_time_s_list": [],
        "loop_copy_mps_wall_time_s_list": [],
        "loop_dmrg_optimization_wall_time_s_list": [],
        "loop_driver_initialize_system_wall_time_s_list": [],
        "loop_generate_initial_mps_wall_time_s_list": [],
        "loop_get_qchem_hami_mpo_wall_time_s_list": [],
        "loop_make_driver_wall_time_s_list": [],
        "loop_reorder_integrals_wall_time_s_list": [],
        "loop_bd1_vscore_list": [],
        "loop_initial_state_vscore_list": [],
        "loop_optket_variance_list": [],
        "loop_initial_ket_energy_list": [],
        "loop_v_score_numerator_list": [],
        "loop_deviation_hf_list": [],
        "loop_deviation_init_ket_list": [],
        "loop_h_min_e_optket_norm_list": [],
        "loop_hf_energy_list": [],
    }
    with h5py.File(data_file, "r") as file_obj:
        # Load for first preloop calculation
        subtract_count = 1
        preloop = file_obj["first_preloop_calc"]["dmrg_results"]
        dmrg_energy = float(preloop["dmrg_ground_state_energy"][()])
        bond_dimension = int(preloop["sweep_bond_dims"][()][-1])
        discarded_weight = float(preloop["sweep_max_discarded_weight"][()][-1])
        loop_cpu_time_s = float(preloop["cpu_single_qchem_dmrg_calc_time_s"][()])
        loop_wall_time_s = float(preloop["wall_single_qchem_dmrg_calc_time_s"][()])

        loop_copy_mps_cpu_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/cpu_copy_mps_time_s"][()]
        )
        loop_dmrg_optimization_cpu_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/cpu_dmrg_optimization_time_s"][()]
        )
        loop_driver_initialize_system_cpu_time_s = float(
            preloop[
                "/first_preloop_calc/dmrg_results/cpu_driver_initialize_system_time_s"
            ][()]
        )
        loop_generate_initial_mps_cpu_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/cpu_generate_initial_mps_time_s"][
                ()
            ]
        )
        loop_get_qchem_hami_mpo_cpu_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/cpu_get_qchem_hami_mpo_time_s"][
                ()
            ]
        )
        loop_make_driver_cpu_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/cpu_make_driver_time_s"][()]
        )
        loop_reorder_integrals_cpu_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/cpu_reorder_integrals_time_s"][()]
        )
        # loop__cpu_time_s = float(preloop["/first_preloop_calc/dmrg_results/cpu_single_qchem_dmrg_calc_time_s"][()])

        loop_copy_mps_wall_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/wall_copy_mps_time_s"][()]
        )
        loop_dmrg_optimization_wall_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/wall_dmrg_optimization_time_s"][
                ()
            ]
        )
        loop_driver_initialize_system_wall_time_s = float(
            preloop[
                "/first_preloop_calc/dmrg_results/wall_driver_initialize_system_time_s"
            ][()]
        )
        loop_generate_initial_mps_wall_time_s = float(
            preloop[
                "/first_preloop_calc/dmrg_results/wall_generate_initial_mps_time_s"
            ][()]
        )
        loop_get_qchem_hami_mpo_wall_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/wall_get_qchem_hami_mpo_time_s"][
                ()
            ]
        )
        loop_make_driver_wall_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/wall_make_driver_time_s"][()]
        )
        loop_reorder_integrals_wall_time_s = float(
            preloop["/first_preloop_calc/dmrg_results/wall_reorder_integrals_time_s"][
                ()
            ]
        )
        if "/first_preloop_calc/dmrg_results/v_score_hartree_fock" in preloop:
            loop_bd1_vscore = float(
                preloop["/first_preloop_calc/dmrg_results/v_score_hartree_fock"][()]
            )
            loop_initial_state_vscore = float(
                preloop["/first_preloop_calc/dmrg_results/v_score_init_ket"][()]
            )
            loop_optket_variance = float(
                preloop["/first_preloop_calc/dmrg_results/optket_variance"][()]
            )
            loop_initial_ket_energy = float(
                preloop["/first_preloop_calc/dmrg_results/initial_ket_energy"][()]
            )
            loop_v_score_numerator = float(
                preloop["/first_preloop_calc/dmrg_results/v_score_numerator"][()]
            )
            loop_deviation_hf = float(
                preloop["/first_preloop_calc/dmrg_results/deviation_hf"][()]
            )
            loop_deviation_init_ket = float(
                preloop["/first_preloop_calc/dmrg_results/deviation_init_ket"][()]
            )
            loop_h_min_e_optket_norm = float(
                preloop["/first_preloop_calc/dmrg_results/h_min_e_optket_norm"][()]
            )
            loop_hf_energy = float(
                preloop["/first_preloop_calc/dmrg_results/hf_energy"][()]
            )

        # loop_single_qchem_dmrg_calc_wall_time_s = float(preloop["/first_preloop_calc/dmrg_results/wall_single_qchem_dmrg_calc_time_s"][()])

        sweep_energies = preloop["sweep_energies"][()].ravel()
        if "reordering_method_used" in preloop:
            reordering_method = preloop["reordering_method_used"][()]
            reordering_method_cpu_times_s.append(
                float(preloop["cpu_reorder_integrals_time_s"][()])
            )
            reordering_method_wall_times_s.append(
                float(preloop["wall_reorder_integrals_time_s"][()])
            )
            reordering_method_list.append(reordering_method)
        else:
            reordering_method_list.append("none (key not found)")

        num_sweeps_list.append(int(len(sweep_energies)))
        final_sweep_delta_energies_list.append(sweep_energies[-1] - sweep_energies[-2])
        dmrg_energies.append(dmrg_energy)
        bond_dimensions.append(bond_dimension)
        discarded_weights.append(discarded_weight)
        loop_cpu_times_s.append(loop_cpu_time_s)
        loop_wall_times_s.append(loop_wall_time_s)
        extra_dict["loop_copy_mps_cpu_time_s_list"].append(loop_copy_mps_cpu_time_s)
        extra_dict["loop_dmrg_optimization_cpu_time_s_list"].append(
            loop_dmrg_optimization_cpu_time_s
        )
        extra_dict["loop_driver_initialize_system_cpu_time_s_list"].append(
            loop_driver_initialize_system_cpu_time_s
        )
        extra_dict["loop_generate_initial_mps_cpu_time_s_list"].append(
            loop_generate_initial_mps_cpu_time_s
        )
        extra_dict["loop_get_qchem_hami_mpo_cpu_time_s_list"].append(
            loop_get_qchem_hami_mpo_cpu_time_s
        )
        extra_dict["loop_make_driver_cpu_time_s_list"].append(
            loop_make_driver_cpu_time_s
        )
        extra_dict["loop_reorder_integrals_cpu_time_s_list"].append(
            loop_reorder_integrals_cpu_time_s
        )

        extra_dict["loop_copy_mps_wall_time_s_list"].append(loop_copy_mps_wall_time_s)
        extra_dict["loop_dmrg_optimization_wall_time_s_list"].append(
            loop_dmrg_optimization_wall_time_s
        )
        extra_dict["loop_driver_initialize_system_wall_time_s_list"].append(
            loop_driver_initialize_system_wall_time_s
        )
        extra_dict["loop_generate_initial_mps_wall_time_s_list"].append(
            loop_generate_initial_mps_wall_time_s
        )
        extra_dict["loop_get_qchem_hami_mpo_wall_time_s_list"].append(
            loop_get_qchem_hami_mpo_wall_time_s
        )
        extra_dict["loop_make_driver_wall_time_s_list"].append(
            loop_make_driver_wall_time_s
        )
        extra_dict["loop_reorder_integrals_wall_time_s_list"].append(
            loop_reorder_integrals_wall_time_s
        )
        if "/first_preloop_calc/dmrg_results/v_score_hartree_fock" in preloop:
            extra_dict["loop_bd1_vscore_list"].append(loop_bd1_vscore)
            extra_dict["loop_initial_state_vscore_list"].append(
                loop_initial_state_vscore
            )
            extra_dict["loop_optket_variance_list"].append(loop_optket_variance)
            extra_dict["loop_initial_ket_energy_list"].append(loop_initial_ket_energy)
            extra_dict["loop_v_score_numerator_list"].append(loop_v_score_numerator)
            extra_dict["loop_deviation_hf_list"].append(loop_deviation_hf)
            extra_dict["loop_deviation_init_ket_list"].append(loop_deviation_init_ket)
            extra_dict["loop_h_min_e_optket_norm_list"].append(loop_h_min_e_optket_norm)
            extra_dict["loop_hf_energy_list"].append(loop_hf_energy)

        # Load for second preloop calculation
        if "second_preloop_calc" in file_obj:
            subtract_count += 1
            preloop = file_obj["second_preloop_calc"]["dmrg_results"]
            dmrg_energy = float(preloop["dmrg_ground_state_energy"][()])
            bond_dimension = int(preloop["sweep_bond_dims"][()][-1])
            discarded_weight = float(preloop["sweep_max_discarded_weight"][()][-1])
            loop_cpu_time_s = float(preloop["cpu_single_qchem_dmrg_calc_time_s"][()])
            loop_wall_time_s = float(preloop["wall_single_qchem_dmrg_calc_time_s"][()])

            loop_copy_mps_cpu_time_s = float(
                preloop["/second_preloop_calc/dmrg_results/cpu_copy_mps_time_s"][()]
            )
            loop_dmrg_optimization_cpu_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/cpu_dmrg_optimization_time_s"
                ][()]
            )
            loop_driver_initialize_system_cpu_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/cpu_driver_initialize_system_time_s"
                ][()]
            )
            loop_generate_initial_mps_cpu_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/cpu_generate_initial_mps_time_s"
                ][()]
            )
            loop_get_qchem_hami_mpo_cpu_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/cpu_get_qchem_hami_mpo_time_s"
                ][()]
            )
            loop_make_driver_cpu_time_s = float(
                preloop["/second_preloop_calc/dmrg_results/cpu_make_driver_time_s"][()]
            )
            loop_reorder_integrals_cpu_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/cpu_reorder_integrals_time_s"
                ][()]
            )
            # loop__cpu_time_s = float(preloop["/second_preloop_calc/dmrg_results/cpu_single_qchem_dmrg_calc_time_s"][()])

            loop_copy_mps_wall_time_s = float(
                preloop["/second_preloop_calc/dmrg_results/wall_copy_mps_time_s"][()]
            )
            loop_dmrg_optimization_wall_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/wall_dmrg_optimization_time_s"
                ][()]
            )
            loop_driver_initialize_system_wall_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/wall_driver_initialize_system_time_s"
                ][()]
            )
            loop_generate_initial_mps_wall_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/wall_generate_initial_mps_time_s"
                ][()]
            )
            loop_get_qchem_hami_mpo_wall_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/wall_get_qchem_hami_mpo_time_s"
                ][()]
            )
            loop_make_driver_wall_time_s = float(
                preloop["/second_preloop_calc/dmrg_results/wall_make_driver_time_s"][()]
            )
            loop_reorder_integrals_wall_time_s = float(
                preloop[
                    "/second_preloop_calc/dmrg_results/wall_reorder_integrals_time_s"
                ][()]
            )
            if "/second_preloop_calc/dmrg_results/v_score_hartree_fock" in preloop:
                loop_bd1_vscore = float(
                    preloop["/second_preloop_calc/dmrg_results/v_score_hartree_fock"][
                        ()
                    ]
                )
                loop_initial_state_vscore = float(
                    preloop["/second_preloop_calc/dmrg_results/v_score_init_ket"][()]
                )
                loop_optket_variance = float(
                    preloop["/second_preloop_calc/dmrg_results/optket_variance"][()]
                )
                loop_initial_ket_energy = float(
                    preloop["/second_preloop_calc/dmrg_results/initial_ket_energy"][()]
                )
                loop_v_score_numerator = float(
                    preloop["/second_preloop_calc/dmrg_results/v_score_numerator"][()]
                )
                loop_deviation_hf = float(
                    preloop["/second_preloop_calc/dmrg_results/deviation_hf"][()]
                )
                loop_deviation_init_ket = float(
                    preloop["/second_preloop_calc/dmrg_results/deviation_init_ket"][()]
                )
                loop_h_min_e_optket_norm = float(
                    preloop["/second_preloop_calc/dmrg_results/h_min_e_optket_norm"][()]
                )
                loop_hf_energy = float(
                    preloop["/second_preloop_calc/dmrg_results/hf_energy"][()]
                )
            # loop_single_qchem_dmrg_calc_wall_time_s = float(preloop["/second_preloop_calc/dmrg_results/wall_single_qchem_dmrg_calc_time_s"][()])

            sweep_energies = preloop["sweep_energies"][()].ravel()
            if "reordering_method_used" in preloop:
                reordering_method = preloop["reordering_method_used"][()]
                reordering_method_cpu_times_s.append(
                    float(preloop["cpu_reorder_integrals_time_s"][()])
                )
                reordering_method_wall_times_s.append(
                    float(preloop["wall_reorder_integrals_time_s"][()])
                )
                reordering_method_list.append(reordering_method)
            else:
                reordering_method_list.append("none (key not found)")

            num_sweeps_list.append(int(len(sweep_energies)))
            final_sweep_delta_energies_list.append(
                sweep_energies[-1] - sweep_energies[-2]
            )
            dmrg_energies.append(dmrg_energy)
            bond_dimensions.append(bond_dimension)
            discarded_weights.append(discarded_weight)
            loop_cpu_times_s.append(loop_cpu_time_s)
            loop_wall_times_s.append(loop_wall_time_s)

            extra_dict["loop_copy_mps_cpu_time_s_list"].append(loop_copy_mps_cpu_time_s)
            extra_dict["loop_dmrg_optimization_cpu_time_s_list"].append(
                loop_dmrg_optimization_cpu_time_s
            )
            extra_dict["loop_driver_initialize_system_cpu_time_s_list"].append(
                loop_driver_initialize_system_cpu_time_s
            )
            extra_dict["loop_generate_initial_mps_cpu_time_s_list"].append(
                loop_generate_initial_mps_cpu_time_s
            )
            extra_dict["loop_get_qchem_hami_mpo_cpu_time_s_list"].append(
                loop_get_qchem_hami_mpo_cpu_time_s
            )
            extra_dict["loop_make_driver_cpu_time_s_list"].append(
                loop_make_driver_cpu_time_s
            )
            extra_dict["loop_reorder_integrals_cpu_time_s_list"].append(
                loop_reorder_integrals_cpu_time_s
            )

            extra_dict["loop_copy_mps_wall_time_s_list"].append(
                loop_copy_mps_wall_time_s
            )
            extra_dict["loop_dmrg_optimization_wall_time_s_list"].append(
                loop_dmrg_optimization_wall_time_s
            )
            extra_dict["loop_driver_initialize_system_wall_time_s_list"].append(
                loop_driver_initialize_system_wall_time_s
            )
            extra_dict["loop_generate_initial_mps_wall_time_s_list"].append(
                loop_generate_initial_mps_wall_time_s
            )
            extra_dict["loop_get_qchem_hami_mpo_wall_time_s_list"].append(
                loop_get_qchem_hami_mpo_wall_time_s
            )
            extra_dict["loop_make_driver_wall_time_s_list"].append(
                loop_make_driver_wall_time_s
            )
            extra_dict["loop_reorder_integrals_wall_time_s_list"].append(
                loop_reorder_integrals_wall_time_s
            )
            if "/second_preloop_calc/dmrg_results/v_score_hartree_fock" in preloop:
                extra_dict["loop_bd1_vscore_list"].append(loop_bd1_vscore)
                extra_dict["loop_initial_state_vscore_list"].append(
                    loop_initial_state_vscore
                )
                extra_dict["loop_optket_variance_list"].append(loop_optket_variance)
                extra_dict["loop_initial_ket_energy_list"].append(
                    loop_initial_ket_energy
                )
                extra_dict["loop_v_score_numerator_list"].append(loop_v_score_numerator)
                extra_dict["loop_deviation_hf_list"].append(loop_deviation_hf)
                extra_dict["loop_deviation_init_ket_list"].append(
                    loop_deviation_init_ket
                )
                extra_dict["loop_h_min_e_optket_norm_list"].append(
                    loop_h_min_e_optket_norm
                )
                extra_dict["loop_hf_energy_list"].append(loop_hf_energy)

        # Load for main loop calculations
        for i in range(1, 1000):
            group_name = f"dmrg_loop_{i:03d}"
            if group_name not in file_obj:
                last_loop = i - 1
                print(f"Last loop included = {last_loop}")

                break
            loop = file_obj[group_name]["dmrg_results"]
            dmrg_energy = float(loop["dmrg_ground_state_energy"][()])
            bond_dimension = int(loop["sweep_bond_dims"][()][-1])
            discarded_weight = float(loop["sweep_max_discarded_weight"][()][-1])
            loop_cpu_time_s = float(loop["cpu_single_qchem_dmrg_calc_time_s"][()])
            loop_wall_time_s = float(loop["wall_single_qchem_dmrg_calc_time_s"][()])

            loop_copy_mps_cpu_time_s = float(loop["cpu_copy_mps_time_s"][()])
            loop_dmrg_optimization_cpu_time_s = float(
                loop["cpu_dmrg_optimization_time_s"][()]
            )
            loop_driver_initialize_system_cpu_time_s = float(
                loop["cpu_driver_initialize_system_time_s"][()]
            )
            loop_generate_initial_mps_cpu_time_s = float(
                loop["cpu_generate_initial_mps_time_s"][()]
            )
            loop_get_qchem_hami_mpo_cpu_time_s = float(
                loop["cpu_get_qchem_hami_mpo_time_s"][()]
            )
            loop_make_driver_cpu_time_s = float(loop["cpu_make_driver_time_s"][()])
            loop_reorder_integrals_cpu_time_s = float(
                loop["cpu_reorder_integrals_time_s"][()]
            )
            # loop__cpu_time_s = float(loop["cpu_single_qchem_dmrg_calc_time_s"][()])

            loop_copy_mps_wall_time_s = float(loop["wall_copy_mps_time_s"][()])
            loop_dmrg_optimization_wall_time_s = float(
                loop["wall_dmrg_optimization_time_s"][()]
            )
            loop_driver_initialize_system_wall_time_s = float(
                loop["wall_driver_initialize_system_time_s"][()]
            )
            loop_generate_initial_mps_wall_time_s = float(
                loop["wall_generate_initial_mps_time_s"][()]
            )
            loop_get_qchem_hami_mpo_wall_time_s = float(
                loop["wall_get_qchem_hami_mpo_time_s"][()]
            )
            loop_make_driver_wall_time_s = float(loop["wall_make_driver_time_s"][()])
            loop_reorder_integrals_wall_time_s = float(
                loop["wall_reorder_integrals_time_s"][()]
            )
            if "v_score_hartree_fock" in loop:
                loop_bd1_vscore = float(loop["v_score_hartree_fock"][()])
                loop_initial_state_vscore = float(loop["v_score_init_ket"][()])
                loop_optket_variance = float(loop["optket_variance"][()])
                loop_initial_ket_energy = float(loop["initial_ket_energy"][()])
                loop_v_score_numerator = float(loop["v_score_numerator"][()])
                loop_deviation_hf = float(loop["deviation_hf"][()])
                loop_deviation_init_ket = float(loop["deviation_init_ket"][()])
                loop_h_min_e_optket_norm = float(loop["h_min_e_optket_norm"][()])
                loop_hf_energy = float(loop["hf_energy"][()])

            # loop_single_qchem_dmrg_calc_wall_time_s = float(loop["wall_single_qchem_dmrg_calc_time_s"][()])

            sweep_energies = loop["sweep_energies"][()].ravel()
            if "reordering_method_used" in loop:
                reordering_method = loop["reordering_method_used"][()]
                reordering_method_cpu_times_s.append(
                    float(loop["cpu_reorder_integrals_time_s"][()])
                )
                reordering_method_wall_times_s.append(
                    float(loop["wall_reorder_integrals_time_s"][()])
                )
                reordering_method_list.append(reordering_method)
            else:
                reordering_method_list.append("none (key not found)")

            num_sweeps_list.append(int(len(sweep_energies)))
            final_sweep_delta_energies_list.append(
                sweep_energies[-1] - sweep_energies[-2]
            )
            dmrg_energies.append(dmrg_energy)
            bond_dimensions.append(bond_dimension)
            discarded_weights.append(discarded_weight)
            loop_cpu_times_s.append(loop_cpu_time_s)
            loop_wall_times_s.append(loop_wall_time_s)

            extra_dict["loop_copy_mps_cpu_time_s_list"].append(loop_copy_mps_cpu_time_s)
            extra_dict["loop_dmrg_optimization_cpu_time_s_list"].append(
                loop_dmrg_optimization_cpu_time_s
            )
            extra_dict["loop_driver_initialize_system_cpu_time_s_list"].append(
                loop_driver_initialize_system_cpu_time_s
            )
            extra_dict["loop_generate_initial_mps_cpu_time_s_list"].append(
                loop_generate_initial_mps_cpu_time_s
            )
            extra_dict["loop_get_qchem_hami_mpo_cpu_time_s_list"].append(
                loop_get_qchem_hami_mpo_cpu_time_s
            )
            extra_dict["loop_make_driver_cpu_time_s_list"].append(
                loop_make_driver_cpu_time_s
            )
            extra_dict["loop_reorder_integrals_cpu_time_s_list"].append(
                loop_reorder_integrals_cpu_time_s
            )

            extra_dict["loop_copy_mps_wall_time_s_list"].append(
                loop_copy_mps_wall_time_s
            )
            extra_dict["loop_dmrg_optimization_wall_time_s_list"].append(
                loop_dmrg_optimization_wall_time_s
            )
            extra_dict["loop_driver_initialize_system_wall_time_s_list"].append(
                loop_driver_initialize_system_wall_time_s
            )
            extra_dict["loop_generate_initial_mps_wall_time_s_list"].append(
                loop_generate_initial_mps_wall_time_s
            )
            extra_dict["loop_get_qchem_hami_mpo_wall_time_s_list"].append(
                loop_get_qchem_hami_mpo_wall_time_s
            )
            extra_dict["loop_make_driver_wall_time_s_list"].append(
                loop_make_driver_wall_time_s
            )
            extra_dict["loop_reorder_integrals_wall_time_s_list"].append(
                loop_reorder_integrals_wall_time_s
            )
            if "v_score_hartree_fock" in loop:
                extra_dict["loop_bd1_vscore_list"].append(loop_bd1_vscore)
                extra_dict["loop_initial_state_vscore_list"].append(
                    loop_initial_state_vscore
                )
                extra_dict["loop_optket_variance_list"].append(loop_optket_variance)
                extra_dict["loop_initial_ket_energy_list"].append(
                    loop_initial_ket_energy
                )
                extra_dict["loop_v_score_numerator_list"].append(loop_v_score_numerator)
                extra_dict["loop_deviation_hf_list"].append(loop_deviation_hf)
                extra_dict["loop_deviation_init_ket_list"].append(
                    loop_deviation_init_ket
                )
                extra_dict["loop_h_min_e_optket_norm_list"].append(
                    loop_h_min_e_optket_norm
                )
                extra_dict["loop_hf_energy_list"].append(loop_hf_energy)

        if "final_dmrg_results" in file_obj:
            print("Processed results available")

            final = file_obj["final_dmrg_results"]
            processed_dmrg_energies = final["past_energies_dmrg"][()]
            processed_bond_dimensions = final["bond_dims_used"][()]
            processed_discarded_weights = final["past_discarded_weights"][()]

            print("Checking that processed results match raw results.")
            assert np.allclose(dmrg_energies, processed_dmrg_energies)
            assert np.allclose(bond_dimensions, processed_bond_dimensions)
            assert np.allclose(discarded_weights, processed_discarded_weights)

    num_loops = last_loop
    num_dmrg_calculations = len(dmrg_energies)
    assert num_loops == num_dmrg_calculations - subtract_count
    return (
        dmrg_energies,
        bond_dimensions,
        discarded_weights,
        num_loops,
        num_dmrg_calculations,
        loop_cpu_times_s,
        loop_wall_times_s,
        num_sweeps_list,
        final_sweep_delta_energies_list,
        reordering_method_list,
        reordering_method_cpu_times_s,
        reordering_method_wall_times_s,
        extra_dict,
    )


def process_time(time_s):
    if time_s > 3600 * 24:
        return f"{time_s / 3600 / 24:.2f} d"
    elif time_s > 3600:
        return f"{time_s / 3600:.2f} h"
    elif time_s > 60:
        return f"{time_s / 60:.2f} m"
    else:
        return f"{time_s:.2f} s"


def add_dmrg_data_chart(
    x_vals_ref,
    y_vals_ref,
    coordinates,
    worksheet,
    x_title="none",
    y_title="none",
    series_name=None,
    chart=None,
    x_min=None,
    x_max=None,
    y_min=None,
    y_max=None,
    marker_symbol="circle",
    marker_fill="156082",
    order=0,
):
    new_chart = False
    if chart is None:
        chart = px_chart.ScatterChart()
        new_chart = True
    chart.style = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title

    series = px_chart.Series(
        values=y_vals_ref,
        xvalues=x_vals_ref,
        title=series_name,
    )
    series.marker.symbol = marker_symbol
    series.marker.graphicalProperties.solidFill = marker_fill  # Marker filling
    series.marker.graphicalProperties.line.solidFill = marker_fill  # Marker outline
    series.marker.graphicalProperties.shadow = "None"  # Marker shadow

    series.graphicalProperties.line.noFill = True

    chart.series.append(series)

    chart.x_axis.scaling.min = x_min
    chart.x_axis.scaling.max = x_max

    chart.y_axis.scaling.min = y_min
    chart.y_axis.scaling.max = y_max

    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "min"

    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.25,
            y=0.01,
            h=0.9,
            w=0.9,
        )
    )

    if series_name is None:
        chart.legend = None

    if new_chart:
        worksheet.add_chart(chart, coordinates)

    return chart


def add_dmrg_processing_basic(
    workbook,
    dmrg_energies,
    bond_dimensions,
    discarded_weights,
    loop_cpu_times_s,
    loop_wall_times_s,
    num_sweeps_list,
    final_sweep_delta_energies_list,
    reordering_method_list,
    reordering_method_cpu_times_s,
    reordering_method_wall_times_s,
    data_dict,
    calc_uuid=None,
    extra_dict=None,
):
    # Get worksheet name from fcidump name
    # take after . and before {
    fcidump_name = data_dict["fcidump"]
    # print(fcidump_name)
    if fcidump_name[:8] == "FCIDUMP_":
        ws_prefix = fcidump_name[8:]
    elif fcidump_name[:8] == "fcidump.":
        ws_prefix = fcidump_name.split(".")[1].split("{")[0]
    else:
        raise ValueError(f"fcidump name not recognized:{fcidump_name}")

    if calc_uuid is not None:
        worksheet_name = ws_prefix + "_" + calc_uuid
    else:
        worksheet_name = ws_prefix
        print("else")
    # print(f"worksheet_name: {worksheet_name}")
    # Create a new worksheet
    ws = workbook.create_sheet(title=worksheet_name)

    # Add header and corresponding data
    ws.append(list(data_dict.keys()))
    ws.append(list(data_dict.values()))

    ws["J3"] = "Estimated Final Energy"
    ws["M3"] = "DON'T USE THIS FOR COMPARISONS, USE E_DMRG INSTEAD (CELL S3)"
    ws["J4"] = "Finish Reason:"
    ws["J5"] = "Loop Results"

    ws["S3"] = "E_DMRG"
    ws["U3"] = (
        "USE THIS E_DMRG FOR COMPARISONS, Estimated Final Energy IS NOT YET RELIABLE"
    )
    ws["S4"] = "DW Linear Slope"
    ws["S5"] = "DW Linear Intercept"

    data_header_dict = {
        "K": "Loop",
        "L": "DMRG Energy",
        "M": "Bond Dimension",
        "N": "Discarded Weights",
        "O": "1/BD",
        "P": "log10 1/BD",
        "Q": "ln DW",
        "R": "Abs Rel Energy",
        "S": "ln AR Energy",
        "T": "Pred E_DMRG",
        "U": "Pred ln AR Energy",
        "V": "CPU Time (s)",
        "W": "Wall Time (s)",
        "X": "CPU Time (processed)",
        "Y": "Wall Time (processed)",
        "Z": "Num Sweeps",
        "AA": "Final Sweep Delta Energy (Eh)",
        "AB": "Reordering Method",
        "AC": "CPU Time Reordering Method (s)",
        "AD": "Wall Time Reordering Method (s)",
    }
    if extra_dict is not None:
        if "loop_bd1_vscore_list" in extra_dict:
            data_header_dict["AE"] = "BD1 V Score"
            data_header_dict["AF"] = "Initial State V Score"
    ws.append(data_header_dict)

    # Fill loop column
    ws["K7"] = -1
    ws["K8"] = 0
    for i in range(len(dmrg_energies) - 2):
        ws[f"K{7+2+i}"] = i + 1

    # Fill DMRG energy column
    for i, energy in enumerate(dmrg_energies):
        ws[f"L{7+i}"] = energy
        ws[f"L{7+i}"].number_format = "0.0000"

    # Fill bond dimension column
    for i, bond_dim in enumerate(bond_dimensions):
        ws[f"M{7+i}"] = bond_dim

    # Fill discarded weights column
    for i, discarded_weight in enumerate(discarded_weights):
        ws[f"N{7+i}"] = discarded_weight
        ws[f"N{7+i}"].number_format = "0.00E+00"

    # Fill 1/BD column with formula
    for i in range(len(bond_dimensions)):
        ws[f"O{7+i}"] = f"=1/M{7+i}"
        ws[f"O{7+i}"].number_format = "0.00E+00"

    # Fill log10 1/BD column with formula
    for i in range(len(bond_dimensions)):
        ws[f"P{7+i}"] = f"=LOG10(O{7+i})"
        ws[f"P{7+i}"].number_format = "0.0000"

    # Fill ln DW column with formula
    for i in range(len(discarded_weights)):
        ws[f"Q{7+i}"] = f"=LN(N{7+i})"
        ws[f"Q{7+i}"].number_format = "0.0000"

    # Fill Abs Rel Energy column with formula
    for i in range(len(dmrg_energies)):
        ws[f"R{7+i}"] = f"=ABS((L{7+i}-$L$3)/$L$3)"
        ws[f"R{7+i}"].number_format = "0.00E+00"

    # Fill ln AR Energy column with formula
    for i in range(len(dmrg_energies)):
        ws[f"S{7+i}"] = f"=LN(R{7+i})"
        ws[f"S{7+i}"].number_format = "0.0000"

    # Fill CPU Time column
    for i, cpu_time in enumerate(loop_cpu_times_s):
        ws[f"V{7+i}"] = cpu_time
        ws[f"V{7+i}"].number_format = "0.00"

    # Fill Wall Time column
    for i, wall_time in enumerate(loop_wall_times_s):
        ws[f"W{7+i}"] = wall_time
        ws[f"W{7+i}"].number_format = "0.00"

    # Fill processed CPU Time column
    for i, cpu_time in enumerate(loop_cpu_times_s):
        ws[f"X{7+i}"] = process_time(cpu_time)

    # Fill processed Wall Time column
    for i, wall_time in enumerate(loop_wall_times_s):
        ws[f"Y{7+i}"] = process_time(wall_time)

    # Fill Num Sweeps column
    for i, num_sweeps in enumerate(num_sweeps_list):
        ws[f"Z{7+i}"] = num_sweeps

    # Fill Final Sweep Delta Energy column
    for i, final_sweep_delta_energy in enumerate(final_sweep_delta_energies_list):
        ws[f"AA{7+i}"] = final_sweep_delta_energy
        ws[f"AA{7+i}"].number_format = "0.00E+00"

    # Fill Reordering Method column
    for i, reordering_method in enumerate(reordering_method_list):
        ws[f"AB{7+i}"] = reordering_method

    # Fill CPU Time column for reordering method
    for i, cpu_time in enumerate(reordering_method_cpu_times_s):
        ws[f"AC{7+i}"] = cpu_time
        ws[f"AC{7+i}"].number_format = "0.00"

    # Fill Wall Time column for reordering method
    for i, wall_time in enumerate(reordering_method_wall_times_s):
        ws[f"AD{7+i}"] = wall_time
        ws[f"AD{7+i}"].number_format = "0.00"

    if extra_dict is not None:
        if "loop_bd1_vscore_list" in extra_dict:
            # Fill BD1 V Score column
            for i, bd1_vscore in enumerate(extra_dict["loop_bd1_vscore_list"]):
                ws[f"AE{7+i}"] = bd1_vscore
                ws[f"AE{7+i}"].number_format = "0.000E+00"

            # Fill Initial State V Score column
            for i, initial_state_vscore in enumerate(
                extra_dict["loop_initial_state_vscore_list"]
            ):
                ws[f"AF{7+i}"] = initial_state_vscore
                ws[f"AF{7+i}"].number_format = "0.000E+00"

    last_data_row = 7 + len(dmrg_energies) - 1
    # E_DMRG formula
    ws["T3"] = f"=MIN(L7:L{last_data_row})"
    ws["T3"].number_format = "0.0000"

    # linear slope and intercept
    ws["T4"] = f"=SLOPE(S7:S{last_data_row},Q7:Q{last_data_row})"
    ws["T4"].number_format = "0.0000"

    ws["T5"] = f"=INTERCEPT(S7:S{last_data_row},Q7:Q{last_data_row})"
    ws["T5"].number_format = "0.0000"

    # Fill Pred E_DMRG column with formula
    for i in range(len(dmrg_energies)):
        ws[f"T{7+i}"] = f"=ABS($L$3)*EXP($T$5)*N{7+i}^$T$4+($L$3)"
        ws[f"T{7+i}"].number_format = "0.0000"

    # Fill Pred ln AR Energy column with formula
    for i in range(len(dmrg_energies)):
        ws[f"U{7+i}"] = f"=$T$4*Q{7+i}+$T$5"
        ws[f"U{7+i}"].number_format = "0.0000"

    # Put in the estimated final energy
    ws["L3"] = dmrg_energies[-1] - 1e-3
    ws["L3"].number_format = "0.000000"

    E_DMRG_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!L7:L{last_data_row}",
    )

    bond_dim_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!M7:M{last_data_row}",
    )

    chart_row_start = 5

    add_dmrg_data_chart(
        x_vals_ref=bond_dim_ref,
        y_vals_ref=E_DMRG_ref,
        coordinates=f"A{chart_row_start}",
        worksheet=ws,
        x_title="Bond Dimension",
        y_title="DMRG Energy (Hartree)",
        series_name=None,
        chart=None,
        x_min=None,
        x_max=None,
        y_min=None,
        y_max=None,
    )

    inv_bond_dim_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!O7:O{last_data_row}",
    )
    chart_row_start += 15
    add_dmrg_data_chart(
        x_vals_ref=inv_bond_dim_ref,
        y_vals_ref=E_DMRG_ref,
        coordinates=f"A{chart_row_start}",
        worksheet=ws,
        x_title="1/(Bond Dimension)",
        y_title="DMRG Energy (Hartree)",
        series_name=None,
        chart=None,
        x_min=None,
        x_max=None,
        y_min=None,
        y_max=None,
    )

    log_inv_bond_dim_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!P7:P{last_data_row}",
    )
    chart_row_start += 15
    add_dmrg_data_chart(
        x_vals_ref=log_inv_bond_dim_ref,
        y_vals_ref=E_DMRG_ref,
        coordinates=f"A{chart_row_start}",
        worksheet=ws,
        x_title="log10(1/(Bond Dimension))",
        y_title="DMRG Energy (Hartree)",
        series_name=None,
        chart=None,
        x_min=None,
        x_max=None,
        y_min=None,
        y_max=None,
    )

    discarded_weight_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!N7:N{last_data_row}",
    )
    E_DMRG_pred_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!T7:T{last_data_row}",
    )
    chart_row_start += 15
    dw_e_dmrg_chart = add_dmrg_data_chart(
        x_vals_ref=discarded_weight_ref,
        y_vals_ref=E_DMRG_pred_ref,
        coordinates=f"A{chart_row_start}",
        worksheet=ws,
        x_title="Discarded Weight",
        y_title="DMRG Energy (Hartree)",
        series_name="Prediction",
        chart=None,
        x_min=None,
        x_max=None,
        y_min=None,
        y_max=None,
        marker_symbol="square",
        marker_fill="00B050",
    )

    add_dmrg_data_chart(
        x_vals_ref=discarded_weight_ref,
        y_vals_ref=E_DMRG_ref,
        coordinates=f"A{chart_row_start}",
        worksheet=ws,
        x_title="Discarded Weight",
        y_title="DMRG Energy (Hartree)",
        series_name="Calculation",
        chart=dw_e_dmrg_chart,
        x_min=None,
        x_max=None,
        y_min=None,
        y_max=None,
    )

    ln_discarded_weight_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!Q7:Q{last_data_row}",
    )
    ln_AR_Energy_pred_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!U7:U{last_data_row}",
    )
    ln_AR_Energy_ref = px.chart.Reference(
        worksheet=ws,
        range_string=f"{ws.title}!S7:S{last_data_row}",
    )
    chart_row_start += 15
    dw_e_dmrg_chart = add_dmrg_data_chart(
        x_vals_ref=ln_discarded_weight_ref,
        y_vals_ref=ln_AR_Energy_pred_ref,
        coordinates=f"A{chart_row_start}",
        worksheet=ws,
        x_title="ln(Discarded Weight)",
        y_title="ln(Abs Relative Energy)",
        series_name="Prediction",
        chart=None,
        x_min=None,
        x_max=None,
        y_min=None,
        y_max=None,
        marker_symbol="square",
        marker_fill="00B050",
    )

    add_dmrg_data_chart(
        x_vals_ref=ln_discarded_weight_ref,
        y_vals_ref=ln_AR_Energy_ref,
        coordinates=f"A{chart_row_start}",
        worksheet=ws,
        x_title="ln(Discarded Weight)",
        y_title="ln(Abs Relative Energy)",
        series_name="Calculation",
        chart=dw_e_dmrg_chart,
        x_min=None,
        x_max=None,
        y_min=None,
        y_max=None,
    )


def setup_workbook(
    data_file_path,
    data_dict_list,
    workbook,
    csv_storage_path="./",
    bd_extrapolation_dict=None,
    memory_summary_csv_filename="./memory_summary.csv",
    csv_uuid=False,
    sweeps_for_normalization=20,
):
    for data_dict in data_dict_list:
        # If data_file_path is a list of paths, try each path until one is found where the file exists
        if isinstance(data_file_path, list):
            for path in data_file_path:

                data_file = (
                    Path(path)
                    / Path(data_dict["Calc UUID"])
                    / Path("dmrg_results.hdf5")
                )
                print(data_file)

                if data_file.exists():
                    break
                else:
                    print(f"File {data_file} does not exist")
                    data_file = (
                        Path(path)
                        / Path(data_dict["fcidump"][8:] + "_" + data_dict["Calc UUID"])
                        / Path("dmrg_results.hdf5")
                    )
                    if data_file.exists():
                        break
        # data_file = (
        #     data_file_path / Path(data_dict["Calc UUID"]) / Path("dmrg_results.hdf5")
        # )
        print(f"Processing {data_file}")
        # Get DMRG energy, bond dimensions, and truncation error for each loop
        (
            dmrg_energies,
            bond_dimensions,
            discarded_weights,
            num_loops,
            num_dmrg_calculations,
            loop_cpu_times_s,
            loop_wall_times_s,
            num_sweeps_list,
            final_sweep_delta_energies_list,
            reordering_method_list,
            reordering_method_cpu_times_s,
            reordering_method_wall_times_s,
            extra_dict,
        ) = get_data_from_incomplete_processing(data_file)

        # Min dmrg energy
        data_dict["Min DMRG Energy"] = min(dmrg_energies)
        # Max bond dimension, dmrg energy at max bond dimension, and computation time at max bond dimension
        bd_max_arg = np.argmax(bond_dimensions)
        data_dict["Max Bond Dimension"] = bond_dimensions[bd_max_arg]
        data_dict["DMRG Energy at Max Bond Dimension"] = dmrg_energies[bd_max_arg]
        data_dict["CPU Time at Max Bond Dimension (hr)"] = (
            loop_cpu_times_s[bd_max_arg] / 3600
        )
        data_dict["Wall Time at Max Bond Dimension (hr)"] = (
            loop_wall_times_s[bd_max_arg] / 3600
        )
        data_dict["num_sweeps_at_max_bd"] = num_sweeps_list[bd_max_arg]

        if "Calc UUID Small BD" in data_dict:
            data_file_small_bd = (
                data_file_path
                / Path(data_dict["Calc UUID Small BD"])
                / Path("dmrg_results.hdf5")
            )
            (
                dmrg_energies_small_bd,
                bond_dimensions_small_bd,
                discarded_weights_small_bd,
                num_loops_small_bd,
                num_dmrg_calculations_small_bd,
                loop_cpu_times_s_small_bd,
                loop_wall_times_s_small_bd,
                num_sweeps_list_small_bd,
                final_sweep_delta_energies_list_small_bd,
                reordering_method_list_small_bd,
                reordering_method_cpu_times_s_small_bd,
                reordering_method_wall_times_s_small_bd,
                extra_dict_small_bd,
            ) = get_data_from_incomplete_processing(data_file_small_bd)
            dmrg_energies = np.hstack((dmrg_energies_small_bd, dmrg_energies))
            bond_dimensions = np.hstack((bond_dimensions_small_bd, bond_dimensions))
            discarded_weights = np.hstack(
                (discarded_weights_small_bd, discarded_weights)
            )
            num_loops = num_loops_small_bd + num_loops
            num_dmrg_calculations = (
                num_dmrg_calculations_small_bd + num_dmrg_calculations
            )
            loop_wall_times_s = np.hstack(
                (loop_wall_times_s_small_bd, loop_wall_times_s)
            )
            loop_cpu_times_s = np.hstack((loop_cpu_times_s_small_bd, loop_cpu_times_s))
            num_sweeps_list = np.hstack((num_sweeps_list_small_bd, num_sweeps_list))
            final_sweep_delta_energies_list = np.hstack(
                (
                    final_sweep_delta_energies_list_small_bd,
                    final_sweep_delta_energies_list,
                )
            )
            reordering_method_list = np.hstack(
                (reordering_method_list_small_bd, reordering_method_list)
            )
            reordering_method_cpu_times_s = np.hstack(
                (reordering_method_cpu_times_s_small_bd, reordering_method_cpu_times_s)
            )
            reordering_method_wall_times_s = np.hstack(
                (
                    reordering_method_wall_times_s_small_bd,
                    reordering_method_wall_times_s,
                )
            )

        add_dmrg_processing_basic(
            workbook=workbook,
            dmrg_energies=dmrg_energies,
            bond_dimensions=bond_dimensions,
            discarded_weights=discarded_weights,
            loop_cpu_times_s=loop_cpu_times_s,
            loop_wall_times_s=loop_wall_times_s,
            num_sweeps_list=num_sweeps_list,
            final_sweep_delta_energies_list=final_sweep_delta_energies_list,
            reordering_method_list=reordering_method_list,
            reordering_method_cpu_times_s=reordering_method_cpu_times_s,
            reordering_method_wall_times_s=reordering_method_wall_times_s,
            data_dict=data_dict,
            calc_uuid=data_dict["Calc UUID"][:4],
            extra_dict=extra_dict,
        )

        # Save performance metrics to csv
        csv_storage_path = Path(csv_storage_path)
        csv_storage_path.mkdir(parents=True, exist_ok=True)
        fcidump_name = data_dict["fcidump"]
        if csv_uuid:
            csv_filename = Path(fcidump_name + "_" + data_dict["Calc UUID"] + ".csv")
        else:
            csv_filename = Path(fcidump_name + ".csv")

        csv_data_array = np.vstack(
            [
                dmrg_energies,
                bond_dimensions,
                discarded_weights,
                loop_cpu_times_s,
                loop_wall_times_s,
            ]
        )
        csv_data_array = csv_data_array.T
        header = (
            "DMRG Energy,Bond Dimension,Discarded Weights,CPU Time (s),Wall Time (s)"
        )

        if extra_dict is not None:
            if "loop_bd1_vscore_list" in extra_dict:
                if len(extra_dict["loop_bd1_vscore_list"]) == len(dmrg_energies):
                    csv_data_array = np.vstack(
                        [
                            csv_data_array.T,
                            extra_dict["loop_bd1_vscore_list"],
                            extra_dict["loop_initial_state_vscore_list"],
                            extra_dict["loop_optket_variance_list"],
                            extra_dict["loop_initial_ket_energy_list"],
                            extra_dict["loop_v_score_numerator_list"],
                            extra_dict["loop_deviation_hf_list"],
                            extra_dict["loop_deviation_init_ket_list"],
                            extra_dict["loop_h_min_e_optket_norm_list"],
                            extra_dict["loop_hf_energy_list"],
                        ]
                    ).T
                    header += ",BD1 V Score,Initial State V Score,Optket Variance,Initial Ket Energy,V Score Numerator,Deviation HF,Deviation Init Ket,H Min E Optket Norm,HF Energy"
        np.savetxt(
            csv_storage_path / csv_filename,
            csv_data_array,
            fmt="%.18e",
            delimiter=",",
            newline="\n",
            header=header,
            footer="",
            comments="# ",
            encoding=None,
        )
    # Save memory summary to csv
    # "fcidump.9_mo_n2-_noncan_0.2_new": {
    #     "energy": -4800.649487589950000,
    #     "energy_95_ci": 0.000172466678630,
    #     "extrapolated_bond_dimension": 325,
    #     "extrapolated_bond_dimension_lower_bound": 304,
    #     "extrapolated_bond_dimension_upper_bound": 349,
    #     "UUID": "54e97bd0-e822-41cc-9c8f-d166194df2b5",
    # },
    if bd_extrapolation_dict is not None:
        extrapolation_df = pd.DataFrame.from_dict(bd_extrapolation_dict, orient="index")
        print(extrapolation_df)
        print(extrapolation_df.columns)
        # energy_series = extrapolation_df["energy"]
        # energy_95_ci_series = extrapolation_df["energy_95_ci"]
        # extrapolated_bond_dimension_series = extrapolation_df["BD Extrapolation"]
        # extrapolated_bond_dimension_lower_bound_series = (
        #     extrapolation_df["BD Extrapolation Lower Bound"]
        # )
        # extrapolated_bond_dimension_upper_bound_series = (
        #     extrapolation_df["BD Extrapolation Upper Bound"]
        # )
        # calc_uuid_series = extrapolation_df["UUID"]

        # bd_extrap_series = extrapolated_bond_dimension_series
        # bd_extrap_series = pd.Series(bd_extrapolation_dict)
        # Join the bd_extrap_series with the data_dict_list
        data_dict_list_df = pd.DataFrame(data_dict_list)
        data_dict_list_df = data_dict_list_df.set_index("fcidump")
        # bd_extrap_series = bd_extrap_series.to_frame()
        # bd_extrap_series.columns = ["BD Extrapolation"]
        data_dict_list_df = data_dict_list_df.join(extrapolation_df)
        data_dict_list_df = data_dict_list_df.reset_index()
        data_dict_list_df = data_dict_list_df.rename(columns={"index": "fcidump"})

        if "total_CPU Time at Max Bond Dimension (hr)" in data_dict_list_df:
            # data_dict_list_df["CPU Time at Max Bond Dimension (hr)"] = (
            #     data_dict_list_df["total_CPU Time at Max Bond Dimension (hr)"]
            # )

            # Do the above replacement, but only if the total_CPU Time at Max Bond Dimension (hr) is not NaN
            data_dict_list_df["CPU Time at Max Bond Dimension (hr)"] = np.where(
                data_dict_list_df["total_CPU Time at Max Bond Dimension (hr)"].isna(),
                data_dict_list_df["CPU Time at Max Bond Dimension (hr)"],
                data_dict_list_df["total_CPU Time at Max Bond Dimension (hr)"],
            )
            data_dict_list_df["total_CPU Time at Max Bond Dimension (hr)"] = data_dict_list_df["CPU Time at Max Bond Dimension (hr)"]
            

        # Use quadratic scaling for the RSS memory usage for extapolated BD, and for CPU and Wall time (cubic)
        data_dict_list_df["RSS Memory Usage (GiB) Extrapolated"] = data_dict_list_df[
            "RSS Memory Usage (GiB)"
        ] * (
            data_dict_list_df["extrapolated_bond_dimension"] ** 2
            / data_dict_list_df["Max Bond Dimension"] ** 2
        )
        data_dict_list_df[
            "CPU Time at Max Bond Dimension (hr) Extrapolated"
        ] = data_dict_list_df["CPU Time at Max Bond Dimension (hr)"] * (
            data_dict_list_df["extrapolated_bond_dimension"] ** 3
            / data_dict_list_df["Max Bond Dimension"] ** 3
        )
        data_dict_list_df[
            "Wall Time at Max Bond Dimension (hr) Extrapolated"
        ] = data_dict_list_df["Wall Time at Max Bond Dimension (hr)"] * (
            data_dict_list_df["extrapolated_bond_dimension"] ** 3
            / data_dict_list_df["Max Bond Dimension"] ** 3
        )
        data_dict_list_df["RSS Memory Usage (GiB) Chem Accuracy"] = np.max(
            [
                data_dict_list_df["RSS Memory Usage (GiB)"],
                data_dict_list_df["RSS Memory Usage (GiB) Extrapolated"],
            ],
            axis=0,
        )

        if "Sweep Normalized CPU Time at Max Bond Dimension (hr)" in data_dict_list_df:
            # For any data that does not already sweep normalized CPU time, calculate it
            data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr)"
                ] = np.where(
                data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr)"
                ].isna(),
                data_dict_list_df["CPU Time at Max Bond Dimension (hr)"]
                / data_dict_list_df["num_sweeps_at_max_bd"]
                * sweeps_for_normalization,
                data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr)"
                ],
            )

            data_dict_list_df[
                "Sweep Normalized CPU Time at Max Bond Dimension (hr) Extrapolated"
            ] = data_dict_list_df[
                "Sweep Normalized CPU Time at Max Bond Dimension (hr)"
            ] * (
                data_dict_list_df["extrapolated_bond_dimension"] ** 3
                / data_dict_list_df["Max Bond Dimension"] ** 3
            )

        # Do the above for the upper and lower bounds on the BD extrapolation

        data_dict_list_df["RSS Memory Usage (GiB) Extrapolated Lower Bound"] = (
            data_dict_list_df["RSS Memory Usage (GiB)"]
            * (
                data_dict_list_df["extrapolated_bond_dimension_lower_bound"]
                / data_dict_list_df["Max Bond Dimension"]
            )
            ** 2
        )

        data_dict_list_df["RSS Memory Usage (GiB) Extrapolated Upper Bound"] = (
            data_dict_list_df["RSS Memory Usage (GiB)"]
            * (
                data_dict_list_df["extrapolated_bond_dimension_upper_bound"]
                / data_dict_list_df["Max Bond Dimension"]
            )
            ** 2
        )

        data_dict_list_df[
            "CPU Time at Max Bond Dimension (hr) Extrapolated Lower Bound"
        ] = (
            data_dict_list_df["CPU Time at Max Bond Dimension (hr)"]
            * (
                data_dict_list_df["extrapolated_bond_dimension_lower_bound"]
                / data_dict_list_df["Max Bond Dimension"]
            )
            ** 3
        )

        data_dict_list_df[
            "CPU Time at Max Bond Dimension (hr) Extrapolated Upper Bound"
        ] = (
            data_dict_list_df["CPU Time at Max Bond Dimension (hr)"]
            * (
                data_dict_list_df["extrapolated_bond_dimension_upper_bound"]
                / data_dict_list_df["Max Bond Dimension"]
            )
            ** 3
        )

        data_dict_list_df[
            "Wall Time at Max Bond Dimension (hr) Extrapolated Lower Bound"
        ] = (
            data_dict_list_df["Wall Time at Max Bond Dimension (hr)"]
            * (
                data_dict_list_df["extrapolated_bond_dimension_lower_bound"]
                / data_dict_list_df["Max Bond Dimension"]
            )
            ** 3
        )

        data_dict_list_df[
            "Wall Time at Max Bond Dimension (hr) Extrapolated Upper Bound"
        ] = (
            data_dict_list_df["Wall Time at Max Bond Dimension (hr)"]
            * (
                data_dict_list_df["extrapolated_bond_dimension_upper_bound"]
                / data_dict_list_df["Max Bond Dimension"]
            )
            ** 3
        )

        data_dict_list_df["RSS Memory Usage (GiB) Chem Accuracy Lower Bound"] = np.max(
            [
                data_dict_list_df["RSS Memory Usage (GiB) Extrapolated Lower Bound"],
                data_dict_list_df["RSS Memory Usage (GiB)"],
            ],
            axis=0,
        )

        data_dict_list_df[
            "RSS Memory Usage (GiB) Chem Accuracy Extrapolated Upper Bound"
        ] = np.max(
            [
                data_dict_list_df["RSS Memory Usage (GiB) Extrapolated Upper Bound"],
                data_dict_list_df["RSS Memory Usage (GiB)"],
            ],
            axis=0,
        )

        

        # Assert that values have correct relative relationships
        # If not, show the values that are incorrect and where the error is
        # Force all rows to be printed
        pd.set_option("display.max_rows", None)
        pd.set_option("display.max_columns", None)
        # print(data_dict_list_df["RSS Memory Usage (GiB) Extrapolated Lower Bound"])
        # print(data_dict_list_df["RSS Memory Usage (GiB) Extrapolated"])
        print(data_dict_list_df)
        print(data_dict_list_df.columns)
        #Remove rows with NaN values
        data_dict_list_df = data_dict_list_df.dropna()

        assert np.all(
            data_dict_list_df["RSS Memory Usage (GiB) Extrapolated Lower Bound"]
            <= data_dict_list_df["RSS Memory Usage (GiB) Extrapolated"] + 1E-3
        ), f"RSS Memory Usage (GiB) Extrapolated Lower Bound not less than RSS Memory Usage (GiB) Extrapolated: {data_dict_list_df['RSS Memory Usage (GiB) Extrapolated Lower Bound'] <= data_dict_list_df['RSS Memory Usage (GiB) Extrapolated']}"


        assert np.all(
            data_dict_list_df["RSS Memory Usage (GiB) Extrapolated"]
            <= data_dict_list_df["RSS Memory Usage (GiB) Extrapolated Upper Bound"] + 1E-3
        )

        assert np.all(
            data_dict_list_df["RSS Memory Usage (GiB)"]
            <= data_dict_list_df[
                "RSS Memory Usage (GiB) Chem Accuracy Extrapolated Upper Bound"
            ] + 1E-3
        )
        assert np.all(
            data_dict_list_df["RSS Memory Usage (GiB) Chem Accuracy"]
            <= data_dict_list_df[
                "RSS Memory Usage (GiB) Chem Accuracy Extrapolated Upper Bound"
            ] + 1E-3
        )

        assert np.all(
            data_dict_list_df[
                "CPU Time at Max Bond Dimension (hr) Extrapolated Lower Bound"
            ]
            <= data_dict_list_df["CPU Time at Max Bond Dimension (hr) Extrapolated"] + 1E-3
        )
        assert np.all(
            data_dict_list_df["CPU Time at Max Bond Dimension (hr) Extrapolated"]
            <= data_dict_list_df[
                "CPU Time at Max Bond Dimension (hr) Extrapolated Upper Bound"
            ] + 1E-3
        )

        assert np.all(
            data_dict_list_df[
                "Wall Time at Max Bond Dimension (hr) Extrapolated Lower Bound"
            ]
            <= data_dict_list_df["Wall Time at Max Bond Dimension (hr) Extrapolated"] + 1E-3
        )
        assert np.all(
            data_dict_list_df["Wall Time at Max Bond Dimension (hr) Extrapolated"]
            <= data_dict_list_df[
                "Wall Time at Max Bond Dimension (hr) Extrapolated Upper Bound"
            ] + 1E-3
        )

        if "Sweep Normalized CPU Time at Max Bond Dimension (hr)" in data_dict_list_df:
            data_dict_list_df[
                "Sweep Normalized CPU Time at Max Bond Dimension (hr) Extrapolated Lower Bound"
            ] = (
                data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr)"
                ]
                * (
                    data_dict_list_df["extrapolated_bond_dimension_lower_bound"]
                    / data_dict_list_df["Max Bond Dimension"]
                )
                ** 3
            )

            data_dict_list_df[
                "Sweep Normalized CPU Time at Max Bond Dimension (hr) Extrapolated Upper Bound"
            ] = (
                data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr)"
                ]
                * (
                    data_dict_list_df["extrapolated_bond_dimension_upper_bound"]
                    / data_dict_list_df["Max Bond Dimension"]
                )
                ** 3
            )

            assert np.all(
                data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr) Extrapolated Lower Bound"
                ]
                <= data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr) Extrapolated"
                ] + 1E-3
            )

            assert np.all(
                data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr) Extrapolated"
                ]
                <= data_dict_list_df[
                    "Sweep Normalized CPU Time at Max Bond Dimension (hr) Extrapolated Upper Bound"
                ] + 1E-3
            )

        data_dict_list_df.to_csv(memory_summary_csv_filename, index=False)


contact_info_temp = [
    {
        "name": "temp",
        "email": "temp",
        "institution": "temp",
    }
]

compute_details_temp = {
    "Machine": "Temp",
    "CPU": "Temp",
    "RAM": "Temp",
}


def produce_set_of_solution_json_files(
    data_file_path,
    data_dict_list,
    json_storage_path="./",
    extrapolation_dict=None,
    instance_dict=None,
    memory_summary_csv_filename="./memory_summary.csv",
    csv_uuid=False,
    contact_info=contact_info_temp,
    compute_details=compute_details_temp,
):
    json_filename_list = []
    for data_dict in data_dict_list:
        # If data_file_path is a list of paths, try each path until one is found where the file exists
        if isinstance(data_file_path, list):
            for path in data_file_path:
                data_file = (
                    Path(path)
                    / Path(data_dict["Calc UUID"])
                    / Path("dmrg_results.hdf5")
                )
                if data_file.exists():
                    break
        (
            dmrg_energies,
            bond_dimensions,
            discarded_weights,
            num_loops,
            num_dmrg_calculations,
            loop_cpu_times_s,
            loop_wall_times_s,
            num_sweeps_list,
            final_sweep_delta_energies_list,
            reordering_method_list,
            reordering_method_cpu_times_s,
            reordering_method_wall_times_s,
            extra_dict,
        ) = get_data_from_incomplete_processing(data_file)

        data_dict["dmrg_energies"] = dmrg_energies
        data_dict["bond_dimensions"] = bond_dimensions
        data_dict["discarded_weights"] = discarded_weights
        data_dict["num_loops"] = num_loops
        data_dict["num_dmrg_calculations"] = num_dmrg_calculations
        data_dict["loop_cpu_times_s"] = loop_cpu_times_s
        data_dict["loop_wall_times_s"] = loop_wall_times_s
        data_dict["num_sweeps_list"] = num_sweeps_list
        data_dict["final_sweep_delta_energies_list"] = final_sweep_delta_energies_list
        data_dict["reordering_method_list"] = reordering_method_list
        data_dict["reordering_method_cpu_times_s"] = reordering_method_cpu_times_s
        data_dict["reordering_method_wall_times_s"] = reordering_method_wall_times_s
        data_dict["extra_dict"] = extra_dict

        # Get extrapolation dict if data_dict["Calc UUID"] in dict
        if extrapolation_dict is not None:
            if data_dict["Calc UUID"] in extrapolation_dict:
                extrapolation_dict_local = extrapolation_dict[data_dict["Calc UUID"]]
            else:
                extrapolation_dict_local = None

        if instance_dict is not None:
            if data_dict["Calc UUID"] in instance_dict:
                instance_dict_local = instance_dict[data_dict["Calc UUID"]]
                problem_instance_uuid = instance_dict_local["problem_instance_uuid"]
                instance_data_object_uuid = instance_dict_local[
                    "instance_data_object_uuid"
                ]
                short_name = instance_dict_local["short_name"]
            else:
                problem_instance_uuid = "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
                instance_data_object_uuid = "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
                short_name = data_dict["fcidump"]

        json_filename = produce_solution_json(
            data_dict,
            contact_info=contact_info,
            compute_hardware_type="classical_computer",
            compute_details=compute_details,
            digital_signature=None,
            json_storage_path=json_storage_path,
            extrapolation_dict=extrapolation_dict_local,
            problem_instance_uuid=problem_instance_uuid,
            instance_data_object_uuid=instance_data_object_uuid,
            short_name=short_name,
        )
        json_filename_list.append(json_filename)

    return json_filename_list


# contact_info_temp_array = [
#     "temp",
#     "temp",
#     "temp",
# ]


def produce_solution_json(
    data_dict,
    contact_info=contact_info_temp,
    compute_hardware_type="classical_computer",
    compute_details=compute_details_temp,
    digital_signature=None,
    json_storage_path="./",
    extrapolation_dict=None,
    problem_instance_uuid="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
    instance_data_object_uuid="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
    short_name="NONE",
):
    # problem_instance_uuid = data_dict["instance ID"]

    # solution_uuid = data_dict["Calc UUID"]
    # short_name = data_dict["fcidump"]

    extra_dict = data_dict["extra_dict"]

    dmrg_energies = data_dict["dmrg_energies"]
    bond_dimensions = data_dict["bond_dimensions"]
    discarded_weights = data_dict["discarded_weights"]

    solution_data = []

    for iiter in range(len(dmrg_energies)):

        overall_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": float(data_dict["loop_wall_times_s"][iiter]),
        }

        preprocessing_sum = (
            float(extra_dict["loop_driver_initialize_system_wall_time_s_list"][iiter])
            + float(extra_dict["loop_generate_initial_mps_wall_time_s_list"][iiter])
            + float(extra_dict["loop_get_qchem_hami_mpo_wall_time_s_list"][iiter])
            + float(extra_dict["loop_reorder_integrals_wall_time_s_list"][iiter])
            + float(extra_dict["loop_make_driver_wall_time_s_list"][iiter])
        )

        preprocessing_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": preprocessing_sum,
        }

        algorithm_run_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": float(
                extra_dict["loop_dmrg_optimization_wall_time_s_list"][iiter]
            ),
        }

        postprocessing_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": float(extra_dict["loop_copy_mps_wall_time_s_list"][iiter]),
        }

        overall_cpu_time = {
            "seconds": float(data_dict["loop_cpu_times_s"][iiter]),
        }

        preprocessing_sum_cpu = (
            float(extra_dict["loop_driver_initialize_system_cpu_time_s_list"][iiter])
            + float(extra_dict["loop_generate_initial_mps_cpu_time_s_list"][iiter])
            + float(extra_dict["loop_get_qchem_hami_mpo_cpu_time_s_list"][iiter])
            + float(extra_dict["loop_reorder_integrals_cpu_time_s_list"][iiter])
            + float(extra_dict["loop_make_driver_cpu_time_s_list"][iiter])
        )

        preprocessing_cpu_time = {
            "seconds": preprocessing_sum_cpu,
        }

        algorithm_run_cpu_time = {
            "seconds": float(
                extra_dict["loop_dmrg_optimization_cpu_time_s_list"][iiter]
            ),
        }

        postprocessing_cpu_time = {
            "seconds": float(extra_dict["loop_copy_mps_cpu_time_s_list"][iiter]),
        }

        time_temp = overall_time

        run_time = {
            "overall_time": overall_time,
            "preprocessing_time": preprocessing_time,
            "algorithm_run_time": algorithm_run_time,
            "postprocessing_time": postprocessing_time,
        }
        run_time_cpu = {
            "overall_time": overall_cpu_time,
            "preprocessing_time": preprocessing_cpu_time,
            "algorithm_run_time": algorithm_run_cpu_time,
            "postprocessing_time": postprocessing_cpu_time,
        }

        solution_data.append(
            {
                "instance_data_object_uuid": instance_data_object_uuid,
                "run_time": run_time,
                "energy": dmrg_energies[iiter],
                "energy_units": "Hartree",
                "run_time_cpu": run_time_cpu,
                "bond_dimension": bond_dimensions[iiter],
                "discarded_weight": discarded_weights[iiter],
                "calculation_uuid": data_dict["Calc UUID"],
            }
        )

        # Add memory info for largest bond dimension
        if iiter == len(dmrg_energies) - 1:
            solution_data[-1]["max_RAM_used_GiB"] = data_dict["RSS Memory Usage (GiB)"]

    # Extrapolated Data
    if extrapolation_dict is not None:
        overall_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": 0.0,
        }

        preprocessing_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": 0.0,
        }

        algorithm_run_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": 0.0,
        }

        postprocessing_time = {
            "wall_clock_start_time": "9999-12-31T23:59:90.999Z",
            "wall_clock_stop_time": "9999-12-31T23:59:99.999Z",
            "seconds": 0.0,
        }

        run_time = {
            "overall_time": overall_time,
            "preprocessing_time": preprocessing_time,
            "algorithm_run_time": algorithm_run_time,
            "postprocessing_time": postprocessing_time,
        }
        solution_data.append(
            {
                "instance_data_object_uuid": instance_data_object_uuid,
                "instance_data_object_short_name": data_dict["fcidump"],
                "run_time": run_time,
                "energy": 0.0,
                "extrapolated_energy": extrapolation_dict["energy"],
                "energy_units": "Hartree",
                "extrapolated_energy_95_ci": extrapolation_dict["energy_95_ci"],
                "extrapolated_bond_dimension": extrapolation_dict[
                    "extrapolated_bond_dimension"
                ],
                "extrapolated_bond_dimension_lower_bound": extrapolation_dict[
                    "extrapolated_bond_dimension_lower_bound"
                ],
                "extrapolated_bond_dimension_upper_bound": extrapolation_dict[
                    "extrapolated_bond_dimension_upper_bound"
                ],
            }
        )

    else:
        pass

    # solution_data = [
    #     problem_instance_uuid,
    #     run_time,
    #     energy,
    #     # "Hartree",
    # ]

    # Check if json file that starts with problem_instance_uuid already exists
    # If not, create a new json file
    # If it does, append to the existing json file
    json_check_path = Path(json_storage_path)
    json_check_path.mkdir(parents=True, exist_ok=True)
    json_files = list(json_check_path.glob(f"{problem_instance_uuid}_*.json"))
    if len(json_files) > 0:

        # Load the existing json file
        with open(json_files[0], "r") as json_file:
            sol_dict = json.load(json_file)

        # Append the new solution data
        sol_dict["solution_data"].extend(solution_data)

        # Add/update last modified timestamp
        last_modified_timestamp = datetime.datetime.now(
            datetime.timezone.utc
        ).isoformat()
        last_modified_timestamp = last_modified_timestamp[:-6] + "Z"
        sol_dict["last_modified_timestamp"] = last_modified_timestamp

        # Save solution to json
        json_filename = json_files[0]
        with open(json_filename, "w") as json_file:
            json.dump(sol_dict, json_file, indent=4)

    else:

        # Timestamp in ISO 8601 format in UTC (note the `Z`) with final Z
        creation_timestamp = datetime.datetime.now(datetime.timezone.utc).isoformat()
        # Replace the time zone shift with Z
        creation_timestamp = creation_timestamp[:-6] + "Z"

        # Generate new solution uuid
        solution_uuid = str(uuid.uuid4())
        sol_dict = {
            "solution_uuid": solution_uuid,
            "problem_instance_uuid": problem_instance_uuid,
            "short_name": short_name,
            "creation_timestamp": creation_timestamp,
            "contact_info": contact_info,
            "solution_data": solution_data,
            "compute_hardware_type": compute_hardware_type,
            "compute_details": compute_details,
            "digital_signature": digital_signature,
            "$schema": "https://github.com/jp7745/qb-file-schemas/blob/main/schemas/solution.schema.0.0.1.json",
        }

        # Save solution to json
        json_filename = Path(json_storage_path) / Path(
            problem_instance_uuid + "_sol_" + solution_uuid + ".json"
        )
        with open(json_filename, "w") as json_file:
            json.dump(sol_dict, json_file, indent=4)

    return json_filename
